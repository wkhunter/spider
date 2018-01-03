#-*- coding: UTF-8 -*-
import sys
import time
import urllib
import urllib.request
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
import chardet

hds = [
    {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
    {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}
]

def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_time = 0
    
    while 1:
        url = 'http://www.douban.com/tag/'+urllib.parse.quote(book_tag)+'/book?start='+str(page_num*15)
        time.sleep(5)
            
        request = urllib.request.Request(url, headers = hds[page_num % len(hds)])
        response = urllib.request.urlopen(request)
        plain_text = response.read()
        # plain_text = source_code

        soup = BeautifulSoup(plain_text, 'lxml')
        list_soup = soup.find('div', {'class': 'mod book-list'})

        if list_soup == None or len(list_soup) <= 1:
            print(len(list_soup))
            break

        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class': 'title'}).string.strip()
            desc = book_info.find('div', {'class': 'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class': 'title'}).get('href')

            try:
                author_info = '/'.join(desc_list[0:3])
            except:
                author_info ='暂无'
            try:
                pub_info = '/'.join(desc_list[-3:])
            except:
                pub_info = '暂无'
            
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating = '0.0'

            book_list.append([title, rating, author_info, pub_info])
            print(str(title) + '---> 下载成功')

        page_num += 1
        print('正在下载第'+ str(page_num) +'页', 'code = ' + str(response.code))

    return book_list

      
def write_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook()
    ws = []
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title = book_tag_lists[i]))
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号','书名','评分', '作者','出版社'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],bl[1],bl[2],bl[3]])
            count += 1
    save_path = 'book_list'
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i])
    save_path += '.xlsx'
    wb.save(save_path)


def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key = lambda x:x[1],reverse = True)
        book_lists.append(book_list)
    return book_lists


if __name__ == '__main__':
    book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教', '心理','判断与决策','算法','数据结构','经济','历史', '传记','哲学','编程','创业','理财','社会学','佛教', '思想','科技','科学','web','股票','爱情','两性', '计算机','机器学习','linux','android','数据库','互联网', '数学', '摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生', '商业','理财','管理', '科普','经典','生活','心灵','文学', '科幻','思维','金融', '名著']
    book_lists = do_spider(book_tag_lists)
    write_book_lists_excel(book_lists, book_tag_lists)

































